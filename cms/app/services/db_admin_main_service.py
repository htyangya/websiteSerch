import json
import os
import sys
from datetime import datetime
import openpyxl
from flask import render_template, current_app, make_response, send_file, Response, flash, session, jsonify
from flask_login import current_user
from openpyxl.styles import PatternFill

import app.lib.cms_lib.session
from app import db
from app.controllers.package import PkgCmsLog
from app.forms.batch_upload_form import BatchUploadForm
from app.forms.privs_dept_form import PrivsDeptForm
from app.forms.privs_user_form import PrivsUserForm
from app.lib.cms_lib.date_util import DateUtil
from app.lib.cms_lib.db_util import DbUtil
from app.lib.cms_lib.str_util import StrUtil
from app.lib.cms_lib.upload_excel_validate_util import UploadExcelValidateUtil
from app.lib.conf.const import Const
from app.models import User
from app.models.cms_db_code_master import CmsDbCodeMaster
from app.models.cms_db_dept_master import CmsDbDeptMaster
from app.models.cms_db_privs_dept import CmsDbPrivsDept
from app.models.cms_db_privs_user import CmsDbPrivsUser
from app.models.cms_object_property import CmsObjectProperty
from app.models.cms_object_type import CmsObjectType
from app.models.cms_operation_log import CmsOperationLog
from app.util.output.daily_log_csv_output import DailyLogCsvOutput


def admin_main_init(db_id, request):
    db_name = ""
    information_message = ""

    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
        information_message = app.lib.cms_lib.session.current_db.information_message
        StrUtil.print_debug('main_db_admin_init. db_name:{0} information_message:{1}'.format(
            db_name, information_message))
    cmsObjectType = CmsObjectType()
    objTypeList = cmsObjectType.getObjectTypeList(db_id)

    return render_template(
        'cms_db_admin/main.html',
        title='ログインメイン',
        db_id=db_id,
        db_name=db_name,
        current_user=current_user,
        objectTypeList=objTypeList,
        appVer=current_app.config['APP_VER'])


# Daily Logメニューの初期化処理
def daily_log_init(db_id):
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name

    # デフォルト値を設定
    form = {}
    form["db_id"] = db_id
    form["log_date1"] = datetime.now().strftime('%Y-%m-%d')
    form["log_date2"] = datetime.now().strftime('%Y-%m-%d')
    form["db_name"] = db_name

    return render_template(
        'cms_db_admin/daily_log.html',
        title='CMS(' + db_name + ') : Daily Log Management',
        current_user=current_user,
        form=form,
        static='init',
        appVer=current_app.config['APP_VER'])


# Daily Logメニューの検索処理
def daily_log_list(form):
    dailyLogList = CmsOperationLog().getCmsDailyLogList(form)

    return render_template(
        'cms_db_admin/daily_log.html',
        title='CMS(' + form['db_name'] + ') : Daily Log Management',
        dailyLogList=dailyLogList,
        current_user=current_user,
        form=form,
        static='search',
        appVer=current_app.config['APP_VER'])


# Daily Logダウンロード
def daily_log_download(form):
    dailyLogList = CmsOperationLog().getCmsDailyLogList(form)

    output = DailyLogCsvOutput()
    output.setCsvTempDir(current_app.config['DOWNLOAD_DIR_PATH'])
    output.setDataList(dailyLogList)

    try:
        output.execute()
    except Exception as e:
        tb = sys.exc_info()[2]
        current_app.logger.error("daily_log_download failed. error_msg:" + str(e.with_traceback(tb)))

    csvFilePath = output.getCsvFileName()
    fileName = 'DailyLogs_' + DateUtil.current_date_str() + '.csv'
    res = make_response(send_file(csvFilePath, attachment_filename=fileName, as_attachment=True))
    res.headers['Content-Type'] = 'application/octet-stream'
    return res


# Privilege (User)
def privs_user_list(db_id):
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
    form = {"add_privs_user": Const.ADD_PRIVS_USER,
            "update_privs_user": Const.UPDATE_PRIVS_USER,
            "delete_privs_user": Const.DELETE_PRIVS_USER,
            "db_id": db_id, "db_name": db_name}

    cmsDbPrivsUser = CmsDbPrivsUser()
    privsUserList = cmsDbPrivsUser.getPrivsUserList(form)

    return render_template(
        'cms_db_admin/privs_user_list.html',
        title='CMS(' + db_name + ') : Privilege User List',
        current_user=current_user,
        privsUserList=privsUserList,
        form=form,
        appVer=current_app.config['APP_VER'])


def privs_user_jqmodal(db_id, request):
    form = PrivsUserForm()
    cmsDbPrivsUser = CmsDbPrivsUser()
    func = request.args.get('func')
    corp_cd = request.args.get('corp_cd')
    dept_cd = request.args.get('dept_cd')
    tuid = request.args.get('tuid')
    privs_type = request.args.get('privs_type')

    privsUser = None
    userDeptList = None
    if func != Const.ADD_PRIVS_USER:
        privsUser = cmsDbPrivsUser.getPrivsUser(db_id, corp_cd, dept_cd, tuid, privs_type)
    if func == Const.UPDATE_PRIVS_USER:
        userDeptList = User.getUserDeptList(tuid)

    return render_template(
        'cms_db_admin/privs_user_jqmodal.html',
        db_id=db_id,
        func=func,
        jqmTitle=Const.PRIVS_USER_TITLE[func],
        form=form,
        privsUser=privsUser,
        userDeptList=userDeptList,
        const=Const)


def get_privs_user_info(request):
    res, msg = {}, "OK"
    user_id = request.form.get('user_id')
    user_info = User.getUserInfo(user_id)
    userDeptList = User.getUserDeptList(user_id)

    if not user_info:
        msg = "User[{}] is not exist.".format(user_id)
    res = {**res, **{"msg": msg}}
    if user_info and user_info.l_lang_family_name != "":
        res = {**res, **{"user_name": user_info.l_lang_family_name + ' ' + user_info.l_lang_first_name}}
    res = {**res, **{"user_dept_list": userDeptList}}

    return Response(json.dumps(res))


def save_privs_user(func, request):
    if len(func) == 0:
        return render_template('error/404.html')

    res, msg = {}, "OK"
    form = PrivsUserForm()
    err_msgs = []
    isSaveError = False

    db_id = request.form["db_id"]
    if func == Const.UPDATE_PRIVS_USER:
        old_corp_cd = request.form["old_corp_cd"]
        old_dept_cd = request.form["old_dept_cd"]
        old_privs_type = request.form["old_privs_type"]
    corp_cd = request.form["corp_cd"]
    dept_cd = request.form["dept_cd"]
    tuid = request.form["user_id"]
    privs_type = request.form["privs_type"]

    cmsDbPrivsUser = CmsDbPrivsUser()

    # 保存処理(新規、編集)
    if func == Const.ADD_PRIVS_USER or func == Const.UPDATE_PRIVS_USER:
        if func == Const.ADD_PRIVS_USER:
            privsUser = cmsDbPrivsUser.getPrivsUser(db_id, corp_cd, dept_cd, tuid, privs_type)
            if privsUser and privsUser.tuid == tuid:
                err_msgs.append(Const.DATA_EXIST_ERR_MSG)
                isSaveError = True
            user_info = User.getUserInfo(tuid)
            if not user_info:
                err_msgs.append(Const.USER_ID_NOT_EXIST_ERR_MSG)
                isSaveError = True
        else:
            privsUser = cmsDbPrivsUser.getPrivsUser(db_id, corp_cd, dept_cd, tuid, privs_type)
            # 登録しようとするデータが存在すれば（自分自身以外）、更新できないよう
            if privsUser and \
                    (privsUser.corp_cd != old_corp_cd
                     or privsUser.dept_cd != old_dept_cd):
                err_msgs.append(Const.DATA_EXIST_ERR_MSG)
                isSaveError = True

        if not isSaveError:
            # 入力チェックする
            cname = [
                "Corp Cd",
                "Department",
                "User Id",
                "Privs Type",
            ]
            input_value = [
                corp_cd,
                dept_cd,
                tuid,
                privs_type,
            ]
            db_field = [
                "MANAGEMENT_CORP_CD",
                "DEPT_CD",
                "TUID",
                "PRIVS_TYPE",
            ]
            col_prop = {'cname': cname, 'input_value': input_value, 'db_field': db_field}
            param_prop = {'err_msgs': [], 'table_name': 'CMS_DB_PRIVS_USER', 'form': form, 'col_prop': col_prop}
            DbUtil.check_input_form_data_by_db(param_prop)

            if len(param_prop['err_msgs']) > 0:
                err_msgs = param_prop['err_msgs']
                isSaveError = True

        if request.method == 'POST' and not isSaveError:
            # form = DatabaseForm(request.form)
            if form.validate_on_submit() == False:
                StrUtil.print_debug("validate error.")
            else:
                try:
                    if func == Const.ADD_PRIVS_USER:
                        addPrivsUser = CmsDbPrivsUser(db_id, corp_cd, dept_cd, tuid, privs_type)
                        cmsDbPrivsUser.addPrivsUser(addPrivsUser, tuid)

                        # Privs User登録を記録する
                        pkgCmsLog = PkgCmsLog()
                        pkgCmsLog.saveOperationLog(
                            current_user.tuid,
                            db_id,
                            operation_cd=Const.OPERATION_CD_ADD_PRIVS_USER,
                            object_id=None,
                            object_type=None,
                            note=tuid
                        )

                        db.session.commit()
                    else:
                        uptPrivsUser = cmsDbPrivsUser.uptPrivsUser(
                            db_id, old_corp_cd, old_dept_cd, tuid, old_privs_type, corp_cd, dept_cd, privs_type,
                            current_user.get_id())

                        # Privs User変更を記録する
                        pkgCmsLog = PkgCmsLog()
                        pkgCmsLog.saveOperationLog(
                            current_user.tuid,
                            db_id,
                            operation_cd=Const.OPERATION_CD_UPDATE_PRIVS_USER,
                            object_id=None,
                            object_type=None,
                            note=tuid
                        )

                        db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    tb = sys.exc_info()[2]
                    StrUtil.print_error("Database save failed. error_msg:{}".format(str(e.with_traceback(tb))))
                    err_msgs.append('Database save failed.')
    # 削除処理
    elif func == Const.DELETE_PRIVS_USER:
        try:
            cmsDbPrivsUser.delPrivsUser(db_id, corp_cd, dept_cd, tuid, privs_type, current_user.get_id())

            # Privs User削除を記録する
            pkgCmsLog = PkgCmsLog()
            pkgCmsLog.saveOperationLog(
                current_user.tuid,
                db_id,
                operation_cd=Const.OPERATION_CD_DELETE_PRIVS_USER,
                object_id=None,
                object_type=None,
                note=tuid
            )

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            tb = sys.exc_info()[2]
            StrUtil.print_error("Database save failed. error_msg:{}".format(str(e.with_traceback(tb))))
            err_msgs.append('Database delete failed.')

    res = {**res, **{"err_msgs": err_msgs}}

    return Response(json.dumps(res))


def privs_dept_list(db_id):
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
    form = {"add_privs_dept": Const.ADD_PRIVS_DEPT,
            "update_privs_dept": Const.UPDATE_PRIVS_DEPT,
            "delete_privs_dept": Const.DELETE_PRIVS_DEPT,
            "db_id": db_id, "db_name": db_name}

    cmsDbPrivsDept = CmsDbPrivsDept()
    privsDeptList = cmsDbPrivsDept.getPrivsDeptList(form)

    return render_template(
        'cms_db_admin/privs_dept_list.html',
        db_id=db_id,
        title='CMS(' + db_name + ') : Privilege Department List',
        current_user=current_user,
        privsDeptList=privsDeptList,
        form=form,
        appVer=current_app.config['APP_VER'])


def privs_dept_jqmodal(db_id, request):
    form = PrivsDeptForm()
    cmsDbPrivsDept = CmsDbPrivsDept()
    func = request.args.get('func')
    corp_cd = request.args.get('corp_cd')
    div_cd = request.args.get('div_cd')
    dept_cd = request.args.get('dept_cd')
    emp_type_cd = request.args.get('emp_type_cd')
    working_type_cd = request.args.get('working_type_cd')
    privs_type = request.args.get('privs_type')

    privsDept = None
    empTypeList = None
    workingTypeList = None
    if func != Const.ADD_PRIVS_DEPT:
        privsDept = cmsDbPrivsDept.getPrivsDept(db_id, corp_cd, div_cd, dept_cd, emp_type_cd, working_type_cd,
                                                privs_type)
    if func != Const.DELETE_PRIVS_DEPT:
        cmsDbCodeMaster = CmsDbCodeMaster()
        empTypeList = cmsDbCodeMaster.getCodeMasterList("EMP_TYPE_CD")
        workingTypeList = cmsDbCodeMaster.getCodeMasterList("WORKING_TYPE_CD")

    return render_template(
        'cms_db_admin/privs_dept_jqmodal.html',
        db_id=db_id,
        func=func,
        jqmTitle=Const.PRIVS_DEPT_TITLE[func],
        form=form,
        privsDept=privsDept,
        empTypeList=empTypeList,
        workingTypeList=workingTypeList,
        const=Const)


def privs_corp_select(db_id, request):
    if len(db_id) == 0:
        return render_template('error/404.html')

    db_name = ""
    result_cnt = 0
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name

    # 組織コードを取得する
    if request.method == 'GET':
        corp_txt = request.args.get('corp_txt')

    if request.method == 'POST':
        corp_txt = request.form['corp_txt']

    if StrUtil.lenb(corp_txt) > 256 and request.method == 'POST':
        flash('検索条件が長すぎます')
        corp_list = None
    else:
        # 組織検索リストを取得
        cmsDbCodeMaster = CmsDbCodeMaster()
        result_cnt = cmsDbCodeMaster.getCorpListCnt(corp_txt)
        corp_list = cmsDbCodeMaster.getCorpList(corp_txt)

        note = corp_txt
        if len(note) > 100:
            note = corp_txt[0:100]

        # 組織検索を記録する
        pkgCmsLog = PkgCmsLog()
        pkgCmsLog.saveOperationLog(
            current_user.tuid,
            db_id,
            operation_cd=Const.OPERATION_CD_CORP_SEARCH,
            note='SearchCond: {}, ResultCnt: {}'.format(note, result_cnt)
        )
        db.session.commit()

    form = {"db_id": db_id, "db_name": db_name, "corp_txt": corp_txt,
            "wait_msg": Const.WAIT_MSG, "select_corp_msg": Const.SELECT_CORP_MSG,
            "zero_list_msg": Const.ZERO_LIST_MSG, }

    return render_template(
        'cms_db_admin/privs_corp_select.html',
        title='Corp Select',
        form=form,
        result_cnt=result_cnt,
        corp_list=corp_list,
        appVer=current_app.config['APP_VER'])


def save_privs_dept(func, request):
    if len(func) == 0:
        return render_template('error/404.html')

    res, msg = {}, "OK"
    form = PrivsDeptForm()
    err_msgs = []
    isSaveError = False

    db_id = request.form["db_id"]
    if func == Const.UPDATE_PRIVS_DEPT:
        old_div_cd = request.form["old_div_cd"]
        old_dept_cd = request.form["old_dept_cd"]
        old_emp_type_cd = request.form["old_emp_type_cd"]
        old_working_type_cd = request.form["old_working_type_cd"]
        old_privs_type = request.form["old_privs_type"]
    corp_cd = request.form["corp_cd"]
    div_cd = request.form["div_cd"]
    dept_cd = request.form["dept_cd"]
    emp_type_cd = request.form["emp_type_cd"]
    working_type_cd = request.form["working_type_cd"]
    privs_type = request.form["privs_type"]

    cmsDbPrivsDept = CmsDbPrivsDept()
    cmsDbCodeMaster = CmsDbCodeMaster()
    OPERATION_NOTE = "CORP_CD={}, DIV_CD={}, DEPT_CD={}, EMP_TYPE_CD={}, WORKING_TYPE_ID={}, PRIVS_TYPE={}"

    # 保存処理(新規、編集)
    if func == Const.ADD_PRIVS_DEPT or func == Const.UPDATE_PRIVS_DEPT:
        if func == Const.ADD_PRIVS_DEPT:
            isCorpCdExist = cmsDbCodeMaster.checkCorpCdExist(corp_cd)
            if not isCorpCdExist:
                err_msgs.append(Const.DATA_NOT_EXIST_ERR_MSG.replace("%s", "Corp Cd"))
                isSaveError = True
            privsDept = cmsDbPrivsDept.getPrivsDept(
                db_id, corp_cd, div_cd, dept_cd, emp_type_cd, working_type_cd, privs_type)
            if privsDept and privsDept.corp_cd == corp_cd:
                err_msgs.append(Const.DATA_EXIST_ERR_MSG)
                isSaveError = True
        else:
            privsDept = cmsDbPrivsDept.getPrivsDept(
                db_id, corp_cd, div_cd, dept_cd, emp_type_cd, working_type_cd, privs_type)
            # 登録しようとするデータが存在すれば（自分自身以外）、更新できないよう
            if privsDept and \
                    (privsDept.div_cd != old_div_cd
                     or privsDept.dept_cd != old_dept_cd
                     or privsDept.emp_type_cd != old_emp_type_cd
                     or privsDept.working_type_cd != old_working_type_cd):
                err_msgs.append(Const.DATA_EXIST_ERR_MSG)
                isSaveError = True

        if not isSaveError:
            # 入力チェックする
            cname = [
                "Corp Cd",
                "Div Cd",
                "Dept Cd",
                "Emp Type",
                "Working Type",
                "Privs Type",
            ]
            input_value = [
                corp_cd,
                div_cd,
                dept_cd,
                emp_type_cd,
                working_type_cd,
                privs_type,
            ]
            db_field = [
                "MANAGEMENT_CORP_CD",
                "DIV_CD",
                "DEPT_CD",
                "EMP_TYPE_CD",
                "WORKING_TYPE_CD",
                "PRIVS_TYPE",
            ]
            col_prop = {'cname': cname, 'input_value': input_value, 'db_field': db_field}
            param_prop = {'err_msgs': [], 'table_name': 'CMS_DB_PRIVS_DEPT', 'form': form, 'col_prop': col_prop}
            DbUtil.check_input_form_data_by_db(param_prop)

            if len(param_prop['err_msgs']) > 0:
                err_msgs = param_prop['err_msgs']
                isSaveError = True

        if request.method == 'POST' and not isSaveError:
            if form.validate_on_submit() == False:
                StrUtil.print_debug("validate error.")
            else:
                try:
                    if func == Const.ADD_PRIVS_DEPT:
                        addPrivsDept = CmsDbPrivsDept(
                            db_id, corp_cd, div_cd, dept_cd, emp_type_cd, working_type_cd, privs_type)
                        cmsDbPrivsDept.addPrivsDept(addPrivsDept, current_user.get_id())

                        # Privs Dept登録を記録する
                        pkgCmsLog = PkgCmsLog()
                        pkgCmsLog.saveOperationLog(
                            current_user.get_id(),
                            db_id,
                            operation_cd=Const.OPERATION_CD_ADD_PRIVS_DEPT,
                            object_id=None,
                            object_type=None,
                            note=OPERATION_NOTE.format(corp_cd, div_cd, dept_cd, emp_type_cd,
                                                       working_type_cd, privs_type)
                        )

                        db.session.commit()
                    else:
                        cmsDbPrivsDept.uptPrivsDept(
                            db_id, corp_cd, div_cd, dept_cd, emp_type_cd, working_type_cd, old_div_cd, old_dept_cd,
                            old_emp_type_cd, old_working_type_cd, old_privs_type, current_user.get_id())

                        # Privs Dept変更を記録する
                        pkgCmsLog = PkgCmsLog()
                        pkgCmsLog.saveOperationLog(
                            current_user.tuid,
                            db_id,
                            operation_cd=Const.OPERATION_CD_UPDATE_PRIVS_DEPT,
                            object_id=None,
                            object_type=None,
                            note=OPERATION_NOTE.format(corp_cd, div_cd, dept_cd, emp_type_cd,
                                                       working_type_cd, privs_type)
                        )

                        db.session.commit()
                except Exception as e:
                    db.session.rollback()
                    tb = sys.exc_info()[2]
                    StrUtil.print_error("Database save failed. error_msg:{}".format(str(e.with_traceback(tb))))
                    err_msgs.append('Database save failed.')
    # 削除処理
    elif func == Const.DELETE_PRIVS_DEPT:
        try:
            cmsDbPrivsDept.delPrivsDept(
                db_id, corp_cd, div_cd, dept_cd, emp_type_cd, working_type_cd, privs_type, current_user.get_id())

            # Privs Dept削除を記録する
            pkgCmsLog = PkgCmsLog()
            pkgCmsLog.saveOperationLog(
                current_user.tuid,
                db_id,
                operation_cd=Const.OPERATION_CD_DELETE_PRIVS_DEPT,
                object_id=None,
                object_type=None,
                note=OPERATION_NOTE.format(corp_cd, div_cd, dept_cd, emp_type_cd, working_type_cd, privs_type)
            )

            db.session.commit()
        except Exception as e:
            db.session.rollback()
            tb = sys.exc_info()[2]
            StrUtil.print_error("Database save failed. error_msg:{}".format(str(e.with_traceback(tb))))
            err_msgs.append('Database delete failed.')

    res = {**res, **{"err_msgs": err_msgs}}

    return Response(json.dumps(res))


def privs_dept_detail(db_id, request):
    corp_cd = request.args.get('corp_cd')
    div_cd = request.args.get('div_cd')
    dept_cd = request.args.get('dept_cd')
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
    form = {"corp_cd": corp_cd,
            "div_cd": div_cd,
            "dept_cd": dept_cd}

    cmsDbDeptMaster = CmsDbDeptMaster()
    deptList = cmsDbDeptMaster.getDeptMasterList(form)

    return render_template(
        'cms_db_admin/privs_dept_detail_list.html',
        db_id=db_id,
        db_name=db_name,
        title='CMS(' + db_name + ') : Privilege Department Detail List',
        current_user=current_user,
        deptList=deptList,
        appVer=current_app.config['APP_VER'])


def batch_upload_get(db_id, request):
    object_type_id = request.args.get('object_type_id')
    object_type_name = CmsObjectType.getCmsObjectType(object_type_id).object_type_name
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
    menu_param = {
        "db_id": db_id,
        "db_name": db_name,
        "object_type_id": object_type_id,
        "object_type_name": object_type_name}

    return render_template(
        'cms_db_admin/object_batch_upload.html',
        db_id=db_id,
        title='CMS(' + db_name + ') : Batch Insert',
        current_user=current_user,
        menu_param=menu_param,
        const=Const,
        appVer=current_app.config['APP_VER'])


def batch_upload_post(db_id, request):
    form = BatchUploadForm()
    if form.ajax_flag.data:
        excel_name = form.save_and_get_filename(False)
        return jsonify({"excel_name": excel_name})
    excel_full_path = form.save_and_get_filename()
    valid = UploadExcelValidateUtil(excel_full_path, db_id, form.object_type_id.data)
    has_error = not valid.validate(form.skip_null_check.data)
    object_type_name = CmsObjectType.getCmsObjectType(form.object_type_id.data).object_type_name
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
    menu_param = {
        "db_name": db_name,
        "object_type_name": object_type_name,
        "has_error": has_error,
        "excel_name": os.path.split(excel_full_path)[-1]
    }
    return render_template(
        'cms_db_admin/object_batch_upload_validate.html',
        db_id=db_id,
        title='CMS(' + db_name + ') : Batch Insert',
        current_user=current_user,
        form=form,
        valid=valid,
        menu_param=menu_param,
        const=Const,
        appVer=current_app.config['APP_VER'])


def upload_data(db_id, request):
    form = BatchUploadForm()
    excel_full_path = form.save_and_get_filename()
    if session.get("more_upload_excel") != excel_full_path:
        valid = UploadExcelValidateUtil(excel_full_path, db_id, form.object_type_id.data)
        item_ctn = valid.save_to_db()
        session["more_upload_excel"] = excel_full_path
        msg = f"Batch Insert was successful.{item_ctn} rows inserted"
    else:
        msg = "Please do not submit data repeatedly"
    object_type_id = form.object_type_id.data
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
    menu_param = {
        "db_name": db_name,
        "object_type_name": CmsObjectType.getCmsObjectType(object_type_id).object_type_name,
        "msg": msg
    }
    return render_template(
        'cms_db_admin/object_batch_upload_success.html',
        db_id=db_id,
        title='CMS(' + db_name + ') : Batch Insert',
        current_user=current_user,
        menu_param=menu_param,
        const=Const,
        appVer=current_app.config['APP_VER'])


# templateダウンロード
def template_download(form):
    object_type_id = form['object_type_id']
    excelTempDir = current_app.config['DOWNLOAD_DIR_PATH']
    cds = DateUtil.current_date_str()
    tmp_path = os.path.join(excelTempDir, cds)
    if not os.path.exists(tmp_path):
        os.makedirs(tmp_path, exist_ok=True)

    fileName = 'CMSObjectList_' + cds + Const.FILE_SUFFIX_EXCEL
    csvFilePath = os.path.join(tmp_path, fileName)
    create_excel(object_type_id, csvFilePath)

    # download file
    res = make_response(send_file(csvFilePath, attachment_filename=fileName, as_attachment=True))
    res.headers['Content-Type'] = 'application/octet-stream'
    return res


def create_excel(object_type_id, csvFilePath):
    db_name = ""
    if app.lib.cms_lib.session.current_db:
        db_name = app.lib.cms_lib.session.current_db.db_name
    data = []
    row2_data = []
    row3_data = []
    row2_data.append('FOLDER NAME')
    row3_data.append('FOLDER')
    cmsObjectProperty = CmsObjectProperty()
    for item in cmsObjectProperty.getObjectPropertyNames(object_type_id):
        row2_data.append(item.property_name)
        row3_data.append(item.property_type)

    data.append(row2_data)
    data.append(row3_data)

    # create excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    ws.cell(row=1, column=1).value = db_name
    fill = PatternFill(patternType='solid', fgColor='b0c4de')

    for info in data:
        ws.append(info)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            ws[cell.coordinate].fill = fill

    # set column width
    adjust_col(ws)
    wb.save(csvFilePath)


def adjust_col(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column  # Get the column name
        column = chr(ord("A") - 1 + column)
        for i in list(ws.rows)[1]:
            try:
                # Necessary to avoid error on empty cells
                if len(str(i.value)) > max_length:
                    max_length = len(i.value)
            except Exception as e:
                pass
        adjusted_width = (max_length + 2) * 1.1
        ws.column_dimensions[column].width = adjusted_width
