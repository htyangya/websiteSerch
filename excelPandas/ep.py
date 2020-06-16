import os

import pandas as pd
import numpy as np


def value_count_str(row):
    vc = (row // 10 * 10).value_counts().sort_index()
    return ",".join(["{0}分数段:{1}人".format(k, v) for k, v in vc.iteritems()])


def get_addr():
    file_name_list = list(filter(lambda x: x.endswith('.xlsx'), os.listdir("./")))
    if not file_name_list:
        return None
    return "./" + file_name_list[0]


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError

    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


df = pd.read_excel(r"C:\Users\tsbcp\Desktop\贵州省2019年省直及垂管系统统一面向社会公开招录公务员（人民警察）总成绩排名.xlsx", 1, 1)
df.columns = df.columns.str.replace(r'\s+', '')
df["成绩区间"] = df.apply(lambda row: int(row["笔试成绩"] / 10) * 10, 1)
ag = df.groupby("报考单位")["笔试成绩"].agg(
    [("最高分", 'max'), ("最低分", 'min'), ("平均分", np.average), ("分数分布", value_count_str)]).round(2)
append_df_to_excel(r"C:\Users\tsbcp\Desktop\respond.xlsx", ag, "sheet2")
